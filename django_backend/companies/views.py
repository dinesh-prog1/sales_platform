from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
import openpyxl, io

from .models import Company
from .serializers import CompanySerializer
from . import services


class CompanyListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        filters = {
            'status': request.query_params.get('status'),
            'size': request.query_params.get('size'),
            'country': request.query_params.get('country'),
            'industry': request.query_params.get('industry'),
            'department': request.query_params.get('department'),
            'search': request.query_params.get('search'),
        }
        companies = services.get_companies(filters)
        serializer = CompanySerializer(companies, many=True)
        return Response({'companies': serializer.data, 'total': companies.count()})

    def post(self, request):
        # Bulk import from JSON list
        data = request.data
        if isinstance(data, list):
            result = services.bulk_create_companies(data)
            return Response(result, status=status.HTTP_201_CREATED)
        # Single create
        serializer = CompanySerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CompanyDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        company = get_object_or_404(Company, pk=pk)
        return Response(CompanySerializer(company).data)

    def patch(self, request, pk):
        company = get_object_or_404(Company, pk=pk)
        serializer = CompanySerializer(company, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        company = get_object_or_404(Company, pk=pk)
        company.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class CompanySearchView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        query = request.query_params.get('q', '')
        companies = services.search_companies(query)
        return Response(CompanySerializer(companies, many=True).data)


class CompanyEmailSuggestionsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        name = request.query_params.get('name', '')
        emails = services.get_email_suggestions(name)
        return Response({'emails': emails})


class CompanyStatusStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(services.get_status_stats())


class CompanySizeStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(services.get_size_stats())


class CompanyUploadExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded'}, status=400)
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        companies = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = dict(zip(headers, row))
            if item.get('email'):
                companies.append({
                    'name': item.get('name') or item.get('Company Name', ''),
                    'email': str(item.get('email') or item.get('Email', '')).lower().strip(),
                    'contact_person': item.get('contact_person') or item.get('Contact Person', ''),
                    'phone': item.get('phone') or item.get('Phone', ''),
                    'website': item.get('website') or item.get('Website', ''),
                    'industry': item.get('industry') or item.get('Industry', ''),
                    'size': item.get('size') or item.get('Size', ''),
                    'country': item.get('country') or item.get('Country', ''),
                    'department': item.get('department') or item.get('Department', ''),
                })
        result = services.bulk_create_companies(companies)
        return Response(result, status=201)
